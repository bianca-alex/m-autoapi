import openpyxl

def write_test_cases_to_excel(test_cases):
    # 创建一个新的Excel工作簿
    workbook = openpyxl.Workbook()
    # 获取默认的工作表
    sheet = workbook.active

    # 设置表头
    headers = ['模块名称', '用例标题', '用例编号', 'url', '请求方法', '请求头', 'cookies', '请求体', '预期结果', '备注']
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)

    # 填写测试用例数据
    for row_num, test_case in enumerate(test_cases, 2):
        sheet.cell(row=row_num, column=1, value=test_case['模块名称'])
        sheet.cell(row=row_num, column=2, value=test_case['用例标题'])
        sheet.cell(row=row_num, column=3, value=test_case['用例编号'])
        sheet.cell(row=row_num, column=4, value=test_case['url'])
        sheet.cell(row=row_num, column=5, value=test_case['请求方法'])
        sheet.cell(row=row_num, column=6, value=test_case['请求头'])
        sheet.cell(row=row_num, column=7, value=test_case['cookies'])
        sheet.cell(row=row_num, column=8, value=test_case['请求体'])
        sheet.cell(row=row_num, column=9, value=test_case['预期结果'])
        sheet.cell(row=row_num, column=10, value=test_case['备注'])

    # 保存Excel文件
    workbook.save('test_cases.xlsx')

# 测试用例数据
test_cases = [
    {
        '模块名称': '用户管理',
        '用例标题': '登录功能验证',
        '用例编号': 'TC001',
        'url': 'http://example.com/login',
        '请求方法': 'POST',
        '请求头': '{"Content-Type": "application/json"}',
        'cookies': '',
        '请求体': '{"username": "testuser", "password": "password123"}',
        '预期结果': '{ "ret": 200, "data": [], "msg": "操作成功" }',
        '备注': ''
    },
    {
        '模块名称': '用户管理',
        '用例标题': '注册功能验证',
        '用例编号': 'TC002',
        'url': 'http://example.com/register',
        '请求方法': 'GET',
        '请求头': '{"Content-Type": "application/json"}',
        'cookies': '',
        '请求体': '',
        '预期结果': '{ "ret": 200, "data": [], "msg": "操作成功" }',
        '备注': ''
    },
    # 添加更多的测试用例数据...
]

if __name__ == "__main__": 
    # 调用函数将测试用例写入Excel文件
    write_test_cases_to_excel(test_cases)
